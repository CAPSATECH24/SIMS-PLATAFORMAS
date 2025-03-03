import openpyxl
import pandas as pd
import sqlite3
import re
import os
import streamlit as st
import logging
from datetime import datetime

# ----------------------------------------------------------------------------- 
# CONFIGURACIÓN INICIAL Y LOGGING 
# ----------------------------------------------------------------------------- 
logging.basicConfig(
    level=logging.INFO, 
    filename='procesamiento.log', 
    filemode='w',
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# ----------------------------------------------------------------------------- 
# BLOQUE 1: FUNCIONES Y LÓGICA PARA DATOS DE PLATAFORMAS 
# (Basado en plataformas_hoy.py) 
# ----------------------------------------------------------------------------- 

default_mappings_plataformas = {
    "WIALON": {
        'Nombre': 'Nombre',
        'Cliente_Cuenta': 'Cuenta',
        'Tipo_de_Dispositivo': 'Tipo de dispositivo',
        'IMEI': 'IMEI',
        'ICCID': 'Iccid',
        'Fecha_de_Activacion': 'Creada',
        'Fecha_de_Desactivacion': 'Desactivación',
        'Hora_de_Ultimo_Mensaje': 'Hora de último mensaje',
        'Ultimo_Reporte': 'Ultimo Reporte',
        'Vehiculo': None,
        'Servicios': None,
        'Grupo': 'Grupos',
        'Telefono': 'Teléfono',
        'Origen': 'WIALON',    # Se asigna manualmente
        'Fecha_Archivo': None  # Se llenará con la fecha detectada en el nombre del archivo
    },
    "ADAS": {
        'Nombre': 'equipo',
        'Cliente_Cuenta': 'Subordinar',
        'Tipo_de_Dispositivo': 'Modelo',
        'IMEI': 'IMEI',
        'ICCID': 'Iccid',
        'Fecha_de_Activacion': 'Activation Date',
        'Fecha_de_Desactivacion': None,
        'Hora_de_Ultimo_Mensaje': None,
        'Ultimo_Reporte': None,
        'Vehiculo': None,
        'Servicios': None,
        'Grupo': None,
        'Telefono': 'Número de tarjeta SIM',
        'Origen': 'ADAS',
        'Fecha_Archivo': None
    },
    "COMBUSTIBLE": {
        'Nombre': 'Vehículo',
        'Cliente_Cuenta': 'Cuenta',
        'Tipo_de_Dispositivo': 'Tanques',
        'IMEI': None,
        'ICCID': None,
        'Fecha_de_Activacion': None,
        'Fecha_de_Desactivacion': None,
        'Hora_de_Ultimo_Mensaje': None,
        'Ultimo_Reporte': 'Último reporte',
        'Vehiculo': 'Vehículo',
        'Servicios': 'Servicios',
        'Grupo': 'Grupos',
        'Telefono': 'Línea',
        'Origen': 'COMBUSTIBLE',
        'Fecha_Archivo': None
    }
}

def create_database_plataformas(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(''' 
        CREATE TABLE IF NOT EXISTS datos ( 
            Nombre TEXT,
            Cliente_Cuenta TEXT,
            Tipo_de_Dispositivo TEXT,
            IMEI TEXT,
            ICCID TEXT,
            Fecha_de_Activacion TEXT,
            Fecha_de_Desactivacion TEXT,
            Hora_de_Ultimo_Mensaje TEXT,
            Ultimo_Reporte TEXT,
            Vehiculo TEXT,
            Servicios TEXT,
            Grupo TEXT,
            Telefono TEXT,
            Origen TEXT,
            Fecha_Archivo TEXT,
            UNIQUE(Nombre, Cliente_Cuenta, Telefono)
        ) 
    ''')
    conn.commit()
    conn.close()

def insert_data_plataformas(db_path, data):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    try:
        cursor.executemany(
            '''INSERT OR IGNORE INTO datos (
                Nombre, Cliente_Cuenta, Tipo_de_Dispositivo, IMEI, ICCID,
                Fecha_de_Activacion, Fecha_de_Desactivacion, Hora_de_Ultimo_Mensaje,
                Ultimo_Reporte, Vehiculo, Servicios, Grupo, Telefono, Origen, Fecha_Archivo
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            data
        )
        conn.commit()
        inserted = cursor.rowcount
        logging.info(f"Insertados {inserted} registros en la base de datos de plataformas.")
    except sqlite3.IntegrityError as e:
        logging.error(f"Error al insertar datos: {e}")
        inserted = 0
    conn.close()
    return inserted

def clean_telefono(telefono):
    if telefono:
        telefono = re.sub(r'\D', '', str(telefono))
        if telefono:
            return telefono
    return None

def extract_date_from_filename(filename):
    match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
    if match:
        return match.group(0)
    else:
        return datetime.now().strftime('%Y-%m-%d')

def process_excel_file_plataformas(excel_file, mappings):
    all_data = []
    invalid_data = []
    total_records = 0

    # Se usa el nombre del archivo subido para extraer la fecha
    filename = excel_file.name
    fecha_archivo = extract_date_from_filename(filename)
    workbook = openpyxl.load_workbook(excel_file, data_only=True)

    for sheet_name in workbook.sheetnames:
        if sheet_name in mappings:
            mapping = mappings[sheet_name]
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            col_indices = {header: idx for idx, header in enumerate(headers) if header is not None}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                total_records += 1
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                record = {}
                is_valid = True

                required_field = 'Cliente_Cuenta'
                column_name = mapping.get(required_field)
                value = row_dict.get(column_name) if column_name else None
                if not value:
                    is_valid = False

                if is_valid:
                    for field in [
                        'Nombre', 'Cliente_Cuenta', 'Tipo_de_Dispositivo', 'IMEI', 'ICCID',
                        'Fecha_de_Activacion', 'Fecha_de_Desactivacion', 'Hora_de_Ultimo_Mensaje',
                        'Ultimo_Reporte', 'Vehiculo', 'Servicios', 'Grupo', 'Telefono', 
                        'Origen', 'Fecha_Archivo'
                    ]:
                        if field == 'Origen':
                            record[field] = mapping['Origen']
                        elif field == 'Fecha_Archivo':
                            record[field] = fecha_archivo
                        else:
                            col_name = mapping.get(field)
                            if col_name:
                                val = row_dict.get(col_name)
                                if field == 'Telefono':
                                    val = clean_telefono(val)
                                record[field] = val
                            else:
                                record[field] = None
                    all_data.append(tuple(record.values()))
                    logging.info(f"Procesado registro válido en '{sheet_name}': {record}")
                else:
                    invalid_data.append(row_dict)
                    logging.warning(f"Registro inválido en '{sheet_name}': {row_dict}")

    return all_data, invalid_data, total_records

# ----------------------------------------------------------------------------- 
# BLOQUE 2: FUNCIONES Y LÓGICA PARA DATOS DE SIMs 
# (Basado en set_sims-companias-unificadas.py) 
# ----------------------------------------------------------------------------- 

default_mappings_sims = {
    "SIMPATIC": {
        'ICCID': 'iccid',
        'TELEFONO': 'msisdn',
        'ESTADO DEL SIM': 'status',
        'EN SESION': 'status',
        'ConsumoMb': 'consumo en Mb'
    },
    "TELCEL ALEJANDRO": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'ESTADO SIM',
        'EN SESION': 'SESIÓN',
        'ConsumoMb': 'LÍMITE DE USO DE DATOS'
    },
    "-1": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado de SIM',
        'EN SESION': 'En sesión',
        'ConsumoMb': 'Uso de ciclo hasta la fecha (MB)'
    },
    "-2": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado de SIM',
        'EN SESION': 'En sesión',
        'ConsumoMb': 'Uso de ciclo hasta la fecha (MB)'
    },
    "TELCEL": {
        'ICCID': 'Cuenta Padre',
        'TELEFONO': 'Línea',
        'ESTADO DEL SIM': 'Estatus línea',
        'EN SESION': 'Estatus línea',
        'ConsumoMb': 'Estatus línea'
    },
    "MOVISTAR": {
        'ICCID': 'ICC',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado',
        'EN SESION': 'Estado GPRS',
        'ConsumoMb': 'Consumo Datos Mensual'
    },
    "NANTI": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado',
        'EN SESION': 'Estado',
        'ConsumoMb': 'Estado'
    },
    "LEGACY": {
        'ICCID': 'ICCID',
        'TELEFONO': 'TELEFONO',
        'ESTADO DEL SIM': 'Estatus',
        'EN SESION': 'Estatus',
        'ConsumoMb': 'BSP Nacional'
    }
}

def create_database_sims(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(''' 
        CREATE TABLE IF NOT EXISTS sims ( 
            ICCID TEXT, 
            TELEFONO TEXT, 
            ESTADO_DEL_SIM TEXT, 
            EN_SESION TEXT, 
            ConsumoMb TEXT,
            Compania TEXT,
            UNIQUE(ICCID, TELEFONO)
        ) 
    ''')
    conn.commit()
    conn.close()

def insert_data_sims(db_path, data):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    records_before = cursor.execute("SELECT COUNT(*) FROM sims").fetchone()[0]
    try:
        cursor.executemany(
            """INSERT OR IGNORE INTO sims (
                ICCID, TELEFONO, ESTADO_DEL_SIM, EN_SESION, ConsumoMb, Compania
            ) VALUES (?, ?, ?, ?, ?, ?)""",
            data
        )
        conn.commit()
        records_after = cursor.execute("SELECT COUNT(*) FROM sims").fetchone()[0]
        records_inserted = records_after - records_before
        logging.info(f"Insertados {records_inserted} registros nuevos en la base de datos de SIMs.")
    finally:
        conn.close()

    return len(data), records_inserted

def clean_iccid_telefono_consumo(data):
    cleaned_data = []
    for row in data:
        cleaned_row = list(row)
        original_iccid = cleaned_row[0]
        original_telefono = cleaned_row[1]
        original_consumo_mb = cleaned_row[4]
        
        if isinstance(original_iccid, float) and original_iccid.is_integer():
            cleaned_iccid = str(int(original_iccid))
        else:
            cleaned_iccid = str(original_iccid)
        cleaned_row[0] = ''.join(filter(str.isdigit, cleaned_iccid)) if cleaned_iccid else ""
        
        if isinstance(original_telefono, float) and original_telefono.is_integer():
            cleaned_telefono = str(int(original_telefono))
        else:
            cleaned_telefono = str(original_telefono)
        cleaned_row[1] = ''.join(filter(str.isdigit, cleaned_telefono)) if cleaned_telefono else ""
        
        if original_consumo_mb:
            cleaned_consumo_mb = ''.join(filter(str.isdigit, str(original_consumo_mb)))
        else:
            cleaned_consumo_mb = ""
        cleaned_row[4] = cleaned_consumo_mb
        
        cleaned_row[2] = cleaned_row[2].strip().lower() if cleaned_row[2] else ""
        cleaned_row[3] = cleaned_row[3].strip().lower() if cleaned_row[3] else ""
        
        cleaned_data.append(tuple(cleaned_row))
        logging.info(
            f"Limpieza Registro SIM: ICCID '{original_iccid}' -> '{cleaned_row[0]}', "
            f"TELEFONO '{original_telefono}' -> '{cleaned_row[1]}', "
            f"ConsumoMb '{original_consumo_mb}' -> '{cleaned_row[4]}'"
        )
    return cleaned_data

def process_excel_sims(excel_file, column_mapping, sheet_name):
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook[sheet_name]
    all_data = []
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = []
        for key in ['ICCID', 'TELEFONO', 'ESTADO DEL SIM', 'EN SESION', 'ConsumoMb']:
            col_index = column_mapping[key]
            if col_index is None or col_index == -1:
                cell_value = ""
            elif col_index >= len(row):
                cell_value = ""
            else:
                cell = row[col_index]
                if isinstance(cell, float) and cell.is_integer():
                    cell_value = str(int(cell))
                elif isinstance(cell, (int, str)):
                    cell_value = str(cell)
                else:
                    cell_value = str(cell) if cell is not None else ""
            row_data.append(cell_value)
        row_data.append(sheet_name)  # Se agrega el nombre de la pestaña como 'Compania'
        all_data.append(row_data)
    return all_data

def process_csv_sims(csv_file, column_mapping):
    try:
        df = pd.read_csv(csv_file, dtype=str)
    except Exception as e:
        logging.error(f"Error leyendo CSV: {e}")
        return []
    
    all_data = []
    company_name = os.path.splitext(os.path.basename(csv_file.name))[0]
    
    for index, row in df.iterrows():
        row_data = []
        for key in ['ICCID', 'TELEFONO', 'ESTADO DEL SIM', 'EN SESION', 'ConsumoMb']:
            col_index = column_mapping[key]
            if col_index is None or col_index == -1:
                cell_value = ""
            else:
                col_name = df.columns[col_index]
                cell = row.get(col_name, "")
                if pd.notnull(cell):
                    cell = cell.strip()
                else:
                    cell = ""
            row_data.append(cell)
        row_data.append(company_name)
        all_data.append(row_data)
    return all_data

# ----------------------------------------------------------------------------- 
# APLICACIÓN STREAMLIT UNIFICADA 
# ----------------------------------------------------------------------------- 

st.title("Aplicación Unificada: Carga de Datos de Plataformas y SIMs")
tabs = st.tabs(["Plataformas", "SIMs"])

# ----------------------------------------------------------------------------- 
# TAB DE PLATAFORMAS 
# ----------------------------------------------------------------------------- 
with tabs[0]:
    st.header("Carga y Homologación de Datos desde Excel (Plataformas)")
    uploaded_file = st.file_uploader("Sube el archivo Excel para Plataformas", type=["xlsx"])
    
    if uploaded_file is not None:
        # Generamos la base de datos en el directorio actual con la fecha de hoy
        today_db_path_plataformas = f"{datetime.now().strftime('%Y-%m-%d')}_plataformas.db"
        if os.path.exists(today_db_path_plataformas):
            st.warning(f"Ya existe una base de datos para hoy (Plataformas): {os.path.basename(today_db_path_plataformas)}")
            if st.button("Eliminar base de datos existente (Plataformas)"):
                try:
                    os.remove(today_db_path_plataformas)
                    st.success("Base de datos de plataformas eliminada correctamente.")
                except Exception as e:
                    st.error(f"Error al eliminar la base de datos de plataformas: {str(e)}")

        if st.button("Ejecutar procesamiento de datos (Plataformas)"):
            all_data, invalid_data, total_records = process_excel_file_plataformas(uploaded_file, default_mappings_plataformas)
            create_database_plataformas(today_db_path_plataformas)

            conn = sqlite3.connect(today_db_path_plataformas)
            cursor = conn.cursor()
            not_inserted = []
            inserted = []
            columns_plat = [
                'Nombre', 'Cliente_Cuenta', 'Tipo_de_Dispositivo', 'IMEI', 'ICCID',
                'Fecha_de_Activacion', 'Fecha_de_Desactivacion', 'Hora_de_Ultimo_Mensaje',
                'Ultimo_Reporte', 'Vehiculo', 'Servicios', 'Grupo', 'Telefono', 'Origen', 'Fecha_Archivo'
            ]
            for record in all_data:
                try:
                    cursor.execute('''
                        INSERT INTO datos (
                            Nombre, Cliente_Cuenta, Tipo_de_Dispositivo, IMEI, ICCID,
                            Fecha_de_Activacion, Fecha_de_Desactivacion, Hora_de_Ultimo_Mensaje,
                            Ultimo_Reporte, Vehiculo, Servicios, Grupo, Telefono, Origen, Fecha_Archivo
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', record)
                    inserted.append(record)
                except sqlite3.IntegrityError:
                    not_inserted.append(record)
            conn.commit()
            conn.close()

            df_inserted = pd.DataFrame(inserted, columns=columns_plat)
            df_not_inserted = pd.DataFrame(not_inserted, columns=columns_plat)

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total de Registros", total_records)
            with col2:
                st.metric("Registros Insertados", len(inserted))
            with col3:
                st.metric("Registros No Insertados", len(not_inserted))
            with col4:
                st.metric("Registros Inválidos", len(invalid_data))

            if len(not_inserted) > 0:
                st.write("### Registros No Insertados (Duplicados)")
                col_a, col_b = st.columns(2)
                with col_a:
                    unique_clients = sorted(df_not_inserted['Cliente_Cuenta'].dropna().unique())
                    selected_client_ni = st.multiselect(
                        'Filtrar por Cliente (No Insertados):',
                        options=unique_clients, 
                        default=[]
                    )
                with col_b:
                    unique_origins = sorted(df_not_inserted['Origen'].dropna().unique())
                    selected_origin_ni = st.multiselect(
                        'Filtrar por Origen (No Insertados):',
                        options=unique_origins, 
                        default=[]
                    )
                df_filtered_ni = df_not_inserted.copy()
                if selected_client_ni:
                    df_filtered_ni = df_filtered_ni[df_filtered_ni['Cliente_Cuenta'].isin(selected_client_ni)]
                if selected_origin_ni:
                    df_filtered_ni = df_filtered_ni[df_filtered_ni['Origen'].isin(selected_origin_ni)]
                st.dataframe(df_filtered_ni, use_container_width=True)
                csv_ni = df_filtered_ni.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Descargar registros no insertados",
                    data=csv_ni,
                    file_name="registros_no_insertados_plataformas.csv",
                    mime='text/csv'
                )

            st.write("## Resumen por Plataforma")
            sheets = list(default_mappings_plataformas.keys())
            summary_data = []
            for sheet in sheets:
                sheet_data = [record for record in all_data if record[-2] == sheet]
                total_sheet = len(sheet_data)
                percentage = (total_sheet / total_records * 100) if total_records > 0 else 0
                summary_data.append({
                    "Plataforma": sheet,
                    "Total Registros": total_sheet,
                    "Porcentaje": f"{percentage:.1f}%"
                })
            df_summary = pd.DataFrame(summary_data)
            st.dataframe(df_summary, use_container_width=True)
            st.write("### Distribución de Registros por Plataforma")
            if not df_summary.empty:
                df_summary['Porcentaje_Num'] = df_summary['Porcentaje'].str.rstrip('%').astype(float)
                chart_data = pd.DataFrame({
                    'Plataforma': df_summary['Plataforma'],
                    'Porcentaje': df_summary['Porcentaje_Num']
                })
                st.bar_chart(chart_data.set_index('Plataforma'))

            platform_tabs = st.tabs(sheets)
            for i, sheet in enumerate(sheets):
                with platform_tabs[i]:
                    st.write(f"## Análisis de {sheet}")
                    sub_data = [record for record in all_data if record[-2] == sheet]
                    total_sheet = len(sub_data)
                    percentage = (total_sheet / total_records * 100) if total_records > 0 else 0
                    st.write("### Resumen de la Plataforma")
                    col_s1, col_s2, col_s3 = st.columns(3)
                    with col_s1:
                        st.metric("Total Registros", total_sheet)
                    with col_s2:
                        st.metric("Porcentaje del Total", f"{percentage:.1f}%")
                    with col_s3:
                        mapped_fields = sum(1 for v in default_mappings_plataformas[sheet].values() if v is not None)
                        st.metric("Campos Mapeados", mapped_fields)
                    if total_sheet > 0:
                        df_sheet = pd.DataFrame(sub_data, columns=columns_plat)
                        st.write("### Datos Filtrables")
                        col_sf1, col_sf2 = st.columns(2)
                        with col_sf1:
                            unique_clients_2 = sorted(df_sheet['Cliente_Cuenta'].dropna().unique())
                            filter_client_2 = st.multiselect(
                                "Filtrar por Cliente:",
                                unique_clients_2, 
                                default=[]
                            )
                        with col_sf2:
                            unique_dev_2 = sorted(df_sheet['Tipo_de_Dispositivo'].dropna().unique())
                            filter_dev_2 = st.multiselect(
                                "Filtrar por Tipo de Dispositivo:",
                                unique_dev_2, 
                                default=[]
                            )
                        df_sheet_filtered = df_sheet.copy()
                        if filter_client_2:
                            df_sheet_filtered = df_sheet_filtered[df_sheet_filtered['Cliente_Cuenta'].isin(filter_client_2)]
                        if filter_dev_2:
                            df_sheet_filtered = df_sheet_filtered[df_sheet_filtered['Tipo_de_Dispositivo'].isin(filter_dev_2)]
                        st.dataframe(df_sheet_filtered, use_container_width=True)
                        csv_sheet = df_sheet_filtered.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label=f"Descargar Datos de {sheet}",
                            data=csv_sheet,
                            file_name=f"{sheet}_datos_plataformas.csv",
                            mime='text/csv'
                        )
                    else:
                        st.warning(f"No hay registros para {sheet}.")

            # --------------------------
            # NUEVA FUNCIONALIDAD: Descargar SQL generado de la base de datos de Plataformas
            # --------------------------
            with sqlite3.connect(today_db_path_plataformas) as conn:
                sql_dump = "\n".join(conn.iterdump())
            st.download_button(
                label="Descargar SQL generado (Plataformas)",
                data=sql_dump,
                file_name=f"{today_db_path_plataformas}.sql",
                mime="text/sql"
            )

# ----------------------------------------------------------------------------- 
# TAB DE SIMs 
# ----------------------------------------------------------------------------- 
with tabs[1]:
    st.header("Carga de Excel/CSV y Homologación de Base de Datos (SIMs)")
    st.write("Sube los archivos Excel o CSV para SIMs")
    
    uploaded_files_sims = st.file_uploader("Selecciona los archivos", type=["xlsx", "csv"], accept_multiple_files=True)
    # Se crea la base de datos de SIMs en el directorio actual
    db_path_sims = "sims_hoy.db"
    
    if uploaded_files_sims:
        # Diccionario para guardar los mapeos (clave = nombre del archivo)
        column_mapping = {}

        for uploaded_file in uploaded_files_sims:
            st.write(f"### Archivo: {uploaded_file.name}")
            if uploaded_file.name.endswith('.xlsx'):
                # Procesamos Excel
                uploaded_file.seek(0)
                workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
                column_mapping[uploaded_file.name] = {}
                for sheet_name in workbook.sheetnames:
                    st.subheader(f"Pestaña: {sheet_name}")
                    header_row = next(workbook[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True))
                    header_row = [col if col else "" for col in header_row]

                    if sheet_name in default_mappings_sims:
                        mapping = default_mappings_sims[sheet_name]
                        mapping_indices = {}
                        mapping_valid = True
                        for key_field, column_name in mapping.items():
                            if column_name in header_row:
                                mapping_indices[key_field] = header_row.index(column_name)
                            else:
                                mapping_valid = False
                                break
                        if mapping_valid:
                            column_mapping[uploaded_file.name][sheet_name] = mapping_indices
                            st.info("Mapeo automático aplicado con éxito.")
                        else:
                            st.warning("No se encontró correspondencia para alguna columna. Selecciona manualmente:")
                            columns_found = header_row
                            iccid_col = st.selectbox("Columna para ICCID:", options=columns_found, key=f"{uploaded_file.name}_{sheet_name}_iccid")
                            telefono_col = st.selectbox("Columna para TELEFONO:", options=columns_found, key=f"{uploaded_file.name}_{sheet_name}_telefono")
                            estado_sim_col = st.selectbox("Columna para ESTADO DEL SIM:", options=columns_found, key=f"{uploaded_file.name}_{sheet_name}_estado")
                            en_sesion_col = st.selectbox("Columna para EN SESION:", options=columns_found, key=f"{uploaded_file.name}_{sheet_name}_sesion")
                            consumo_mb_col = st.selectbox("Columna para ConsumoMb:", options=columns_found, key=f"{uploaded_file.name}_{sheet_name}_consumo")
                            column_mapping[uploaded_file.name][sheet_name] = {
                                'ICCID': columns_found.index(iccid_col),
                                'TELEFONO': columns_found.index(telefono_col),
                                'ESTADO DEL SIM': columns_found.index(estado_sim_col),
                                'EN SESION': columns_found.index(en_sesion_col),
                                'ConsumoMb': columns_found.index(consumo_mb_col)
                            }
                    else:
                        st.info("Pestaña no definida en mapeo por defecto. Selecciona manualmente:")
                        header_row = next(workbook[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True))
                        header_row = [col if col else "" for col in header_row]
                        iccid_col = st.selectbox("Columna para ICCID:", options=header_row, key=f"{uploaded_file.name}_{sheet_name}_iccid_man")
                        telefono_col = st.selectbox("Columna para TELEFONO:", options=header_row, key=f"{uploaded_file.name}_{sheet_name}_tel_man")
                        estado_sim_col = st.selectbox("Columna para ESTADO DEL SIM:", options=header_row, key=f"{uploaded_file.name}_{sheet_name}_estado_man")
                        en_sesion_col = st.selectbox("Columna para EN SESION:", options=header_row, key=f"{uploaded_file.name}_{sheet_name}_sesion_man")
                        consumo_mb_col = st.selectbox("Columna para ConsumoMb:", options=header_row, key=f"{uploaded_file.name}_{sheet_name}_consumo_man")
                        column_mapping[uploaded_file.name][sheet_name] = {
                            'ICCID': header_row.index(iccid_col),
                            'TELEFONO': header_row.index(telefono_col),
                            'ESTADO DEL SIM': header_row.index(estado_sim_col),
                            'EN SESION': header_row.index(en_sesion_col),
                            'ConsumoMb': header_row.index(consumo_mb_col)
                        }
            elif uploaded_file.name.endswith('.csv'):
                st.subheader("Archivo CSV")
                try:
                    df_csv = pd.read_csv(uploaded_file, dtype=str)
                except Exception as e:
                    st.error(f"Error leyendo CSV: {e}")
                    continue
                columns_csv = df_csv.columns.tolist()
                file_name_no_ext = os.path.splitext(uploaded_file.name)[0]
                if file_name_no_ext in default_mappings_sims:
                    mapping = default_mappings_sims[file_name_no_ext]
                    mapping_indices = {}
                    mapping_valid = True
                    for key_field, column_name in mapping.items():
                        if column_name in columns_csv:
                            mapping_indices[key_field] = columns_csv.index(column_name)
                        else:
                            mapping_valid = False
                            break
                    if mapping_valid:
                        column_mapping[uploaded_file.name] = mapping_indices
                        st.info("Mapeo automático aplicado con éxito para CSV.")
                    else:
                        st.warning("Algunas columnas no se encontraron. Selección manual:")
                        iccid_col = st.selectbox("Columna para ICCID:", options=columns_csv, key=f"{uploaded_file.name}_iccid_man")
                        telefono_col = st.selectbox("Columna para TELEFONO:", options=columns_csv, key=f"{uploaded_file.name}_tel_man")
                        estado_sim_col = st.selectbox("Columna para ESTADO DEL SIM:", options=columns_csv, key=f"{uploaded_file.name}_estado_man")
                        en_sesion_col = st.selectbox("Columna para EN SESION:", options=columns_csv, key=f"{uploaded_file.name}_sesion_man")
                        consumo_mb_col = st.selectbox("Columna para ConsumoMb:", options=columns_csv, key=f"{uploaded_file.name}_consumo_man")
                        column_mapping[uploaded_file.name] = {
                            'ICCID': columns_csv.index(iccid_col),
                            'TELEFONO': columns_csv.index(telefono_col),
                            'ESTADO DEL SIM': columns_csv.index(estado_sim_col),
                            'EN SESION': columns_csv.index(en_sesion_col),
                            'ConsumoMb': columns_csv.index(consumo_mb_col)
                        }
                else:
                    st.info("CSV sin mapeo predefinido. Selección manual de columnas:")
                    iccid_col = st.selectbox("Columna para ICCID:", options=columns_csv, key=f"{uploaded_file.name}_iccid_man2")
                    telefono_col = st.selectbox("Columna para TELEFONO:", options=columns_csv, key=f"{uploaded_file.name}_tel_man2")
                    estado_sim_col = st.selectbox("Columna para ESTADO DEL SIM:", options=columns_csv, key=f"{uploaded_file.name}_estado_man2")
                    en_sesion_col = st.selectbox("Columna para EN SESION:", options=columns_csv, key=f"{uploaded_file.name}_sesion_man2")
                    consumo_mb_col = st.selectbox("Columna para ConsumoMb:", options=columns_csv, key=f"{uploaded_file.name}_consumo_man2")
                    column_mapping[uploaded_file.name] = {
                        'ICCID': columns_csv.index(iccid_col),
                        'TELEFONO': columns_csv.index(telefono_col),
                        'ESTADO DEL SIM': columns_csv.index(estado_sim_col),
                        'EN SESION': columns_csv.index(en_sesion_col),
                        'ConsumoMb': columns_csv.index(consumo_mb_col)
                    }

        if st.button("Procesar Archivos de SIMs"):
            create_database_sims(db_path_sims)
            logging.info(f"Base de datos de SIMs creada/verificada: {db_path_sims}")

            total_records_sims = 0
            total_inserted_sims = 0
            stats_by_file = {}

            for uploaded_file in uploaded_files_sims:
                if uploaded_file.name.endswith('.xlsx'):
                    uploaded_file.seek(0)
                    workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
                    stats_by_file[uploaded_file.name] = {'sheets': {}}
                    for sheet_name in column_mapping[uploaded_file.name].keys():
                        # Es importante reiniciar el puntero en el archivo para cada lectura
                        uploaded_file.seek(0)
                        data = process_excel_sims(uploaded_file, column_mapping[uploaded_file.name][sheet_name], sheet_name)
                        data_cleaned = clean_iccid_telefono_consumo(data)
                        processed, inserted = insert_data_sims(db_path_sims, data_cleaned)
                        stats_by_file[uploaded_file.name]['sheets'][sheet_name] = {
                            'processed': processed,
                            'inserted': inserted
                        }
                        total_records_sims += processed
                        total_inserted_sims += inserted
                elif uploaded_file.name.endswith('.csv'):
                    uploaded_file.seek(0)
                    data = process_csv_sims(uploaded_file, column_mapping[uploaded_file.name])
                    data_cleaned = clean_iccid_telefono_consumo(data)
                    processed, inserted = insert_data_sims(db_path_sims, data_cleaned)
                    stats_by_file[uploaded_file.name] = {
                        'processed': processed,
                        'inserted': inserted
                    }
                    total_records_sims += processed
                    total_inserted_sims += inserted

            st.success("¡Procesamiento de SIMs completado!")
            st.write(f"Total de registros procesados: {total_records_sims}")
            st.write(f"Total de registros insertados (evitando duplicados): {total_inserted_sims}")

            st.write("### Estadísticas de Procesamiento por Archivo/Pestaña")
            for file, info in stats_by_file.items():
                st.subheader(f"Archivo: {file}")
                if 'sheets' in info:
                    for sheet, sheet_stats in info['sheets'].items():
                        processed = sheet_stats['processed']
                        inserted = sheet_stats['inserted']
                        insertion_rate = (inserted / processed * 100) if processed else 0
                        st.write(f"**Pestaña:** {sheet}")
                        col_a1, col_a2, col_a3 = st.columns(3)
                        with col_a1:
                            st.metric("Registros Procesados", processed)
                        with col_a2:
                            st.metric("Registros Insertados", inserted)
                        with col_a3:
                            st.metric("Tasa de Inserción", f"{insertion_rate:.2f}%")
                else:
                    processed = info['processed']
                    inserted = info['inserted']
                    insertion_rate = (inserted / processed * 100) if processed else 0
                    col_b1, col_b2, col_b3 = st.columns(3)
                    with col_b1:
                        st.metric("Registros Procesados", processed)
                    with col_b2:
                        st.metric("Registros Insertados", inserted)
                    with col_b3:
                        st.metric("Tasa de Inserción", f"{insertion_rate:.2f}%")
            
            # --------------------------
            # NUEVA FUNCIONALIDAD: Descargar SQL generado de la base de datos de SIMs
            # --------------------------
            with sqlite3.connect(db_path_sims) as conn:
                sql_dump_sims = "\n".join(conn.iterdump())
            st.download_button(
                label="Descargar SQL generado (SIMs)",
                data=sql_dump_sims,
                file_name="sims_hoy.db.sql",
                mime="text/sql"
            )
    else:
        st.warning("No se han subido archivos para SIMs.")
